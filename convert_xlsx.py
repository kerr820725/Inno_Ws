#In[]
from collections import defaultdict
import math
import socket , glob ,datetime , configparser
import pandas as pd
from openpyxl import load_workbook
from functools import reduce
import json


class Xlsx:
    def __init__(self, csv_filter_name= 'qa' ):#csv_filename_path csv 的路徑檔名
        #self.pc_name = socket.gethostname().split('-')[1]
        
        config = configparser.ConfigParser(comment_prefixes='/', allow_no_value=True)
        config.read('config.ini', encoding="utf-8")
        
        
        self.csv_filename = config['config']['csv_filename_path'] +  '/{csv_filename}*.csv'.format(csv_filename = csv_filter_name )

        self.csv_list = glob.glob(self.csv_filename )# 找尋所有 相關 檔案名 , 回傳list
        self.xlsx_filename = '%s/%s.xlsx'%( config['config']['csv_filename_path']  ,datetime.datetime.now().strftime('%Y%m%d%H%M%S'))#時間檔名的  xlsx
        self.test_type = config['paltform_rain_ws']['test_type']
        self.goal_time = config['paltform_rain_ws']['goal_event_time']
    
    def csv_to_xlsx(self, csv_name ):# Csv 轉 Xlsx 
        csv = pd.read_csv('{csv_name}'.format(csv_name=csv_name)   , encoding='latin1' ,error_bad_lines=False  )
        
        csv.to_excel( self.xlsx_filename  ,
        sheet_name= 'all' , index = False, header= True)

        print('csv conver to xlsx Ok , filename : %s '%self.xlsx_filename)

        self.excel_book = load_workbook(self.xlsx_filename )
    
    def update_sheet_1(self , csv_name):#同個 sheet update 其他 csv檔 用法

        with pd.ExcelWriter(self.xlsx_filename, engine='openpyxl') as writer:

            writer.book =  self.excel_book

            writer.sheets = {ws.title: ws for ws in  self.excel_book.worksheets}

            csv = pd.read_csv('{csv_name}'.format(csv_name= csv_name )  
                , encoding='latin1' , error_bad_lines=False)

            # 同個 sheet update 主要是用 startrow 參數
            csv.to_excel(writer, 'all',  index = False,header= False ,
                startrow =writer.sheets['all'].max_row)
            
            print('%s update 成功 writer.sheets: %s'%(csv_name,writer.sheets))

            writer.save()

    def add_column_to_sheet_1(self): #增加goal_event_time欄位
        df = pd.read_excel(self.xlsx_filename, sheet_name = 'all', engine='openpyxl', header=0)
        with pd.ExcelWriter(self.xlsx_filename, engine='openpyxl') as writer:
            writer.book = self.excel_book
            writer.sheets = {ws.title: ws for ws in  self.excel_book.worksheets}
            
            #原始進球時間戳
            goal_times = [i for i in self.goal_time.split(',')]
            #去除毫秒，與start_time比較用
            goal_times_compare = [int(i[:-3] + '000') for i in self.goal_time.split(',')]
            #存放goal_event_time欄位的資料
            goals = []
            for  index in df.index:
                action = df.iloc[index]['action']
                # 若action不是rain，取start_time
                if action not in ['Rain']:
                    start_time = df.iloc[index]['start_time']
                    goals.append(start_time)
                    continue

                start_time = int(df.iloc[index]['start_time']) 
                for i in range(len(goal_times_compare)-1, -1 , -1):
                    #start_time比當前進球時間大，取當前進球時間
                    if start_time >= goal_times_compare[i]:
                        goals.append(goal_times[i])
                        break

            df['goal_event_time'] = goals
            df.to_excel(writer, 'all' , index= False , header=True)
            writer.save()
            print('新增 %s 欄位成功 writer.sheets: %s'%('goal_event_time' , writer.sheets))

    def add_sheet_data(self, sheet_name ): # 新增 sheet ,並增加 相對應 data 邏輯 
        
        with pd.ExcelWriter(self.xlsx_filename, engine='openpyxl') as writer:

            writer.book = self.excel_book

            frame_data = self.generate_data(sheet_name=sheet_name)
            secondMockDF = pd.DataFrame(frame_data)

            if sheet_name == 'statistic':
                
                secondMockDF = secondMockDF.set_index(('動作'))

                # 這邊移除 後 會讓 excel 多出一行空白 ,但不做 會 顯示 兩個動作
                secondMockDF.index.name = ''

                cols = [  (self.all_time_title , '次數'), (self.all_time_title , '成功'), (self.all_time_title , '成功率') 
                    , (self.all_time_title , 'Avg.Response Time'), (self.all_time_title , 'Throughput(APIs / sec)')]

                secondMockDF.columns = pd.MultiIndex.from_tuples(cols , names= ['' , '動作' ])

                print (secondMockDF)
                secondMockDF.to_excel(writer, sheet_name , index= True , header=True)
            else: # 另外個 延遲時間 不需index
                secondMockDF.to_excel(writer, sheet_name , index= False , header=True)
                 
            writer.sheets = {ws.title: ws for ws in self.excel_book.worksheets }
            print('新增 %s 成功 writer.sheets: %s'%(sheet_name , writer.sheets))

            # Save the file
            #writer.save()

    def xlsx_data_transfer(self):# 這裡是 讀取 xlsx sheet all 後 ,並轉成需要的格式
        self.statistic_dict = {}
        self.socket_rain_success = []# 用來將 rain 成功 狀態 0 放進來 , sheet 3 統計 時間 區段用
        self.all_user = []
        self.rain_user_list = defaultdict(list)
        
        df = pd.read_excel(self.xlsx_filename ).to_dict('records')
        init_time = ''

        for  index , df_dict in enumerate(df):
            try:
                break_flag = ''           
                action = df_dict['action'].split('?')[0]
                if action == 'Rain' and self.test_type == 'goal':
                    #進球紅包雨

                    goal_event_time = int(df_dict['goal_event_time'])
                    end_time =  int(df_dict['end_time'].split('.')[0])
                    diff =  float(end_time - goal_event_time ) /1000
                    dict_json = json.loads(df_dict['response'])
                    masterId = dict_json['data']['masterId']

                else:
                    diff = float(df_dict['diff'] )
                    if 'user/token' in action:
                        self.all_user.append(df_dict['user'] )
                    if action == 'Rain':
                        dict_json = json.loads(df_dict['response'])
                        masterId = dict_json['data']['masterId']
                        
                        
              
                try:
                    status_code =  str(df_dict['status_code'] ).split('.')[0]#  會有 200.0  這種
                except:
                    status_code = str(df_dict['status_code'] )

                if 'WS_Closed' in action or 'OtherWSMsg' in action: # diff 有問題  ,先過濾掉
                    continue

                if 'WS_Error' in action :
                    new_action = 'WS_Error: %s'%df_dict['response']
                    if new_action not in  self.statistic_dict.keys() :
                        self.statistic_dict[ new_action ] = {'error_times' : [] }
                    self.statistic_dict[ new_action ]['error_times'].append( 1 )
                else:# 其他action 
                    if action not in  self.statistic_dict.keys() :
                        self.statistic_dict[ action ] = {'status_code': [] , 'diff': [] }
                        
                    self.statistic_dict[ action ]['status_code'].append( status_code )
                    
                    if not math.isnan(diff):
                        self.statistic_dict[ action]['diff'].append( diff )
                    

   
                    if action == 'Rain' :
                    
                        if status_code == '0' :# 將 rain 成功的放到  list , 給 socket 統計居間用
                        
                            if  df_dict['user'] not in self.rain_user_list[masterId]: # 同個master id  如果遇到 相同user 的rain , 只會拿一次
                                self.socket_rain_success.append(diff)

                        self.rain_user_list[masterId].append( df_dict['user']  )

                if 'Rain' not in action:
                    if init_time == '':
                        try:
                            init_time = datetime.datetime.strptime(df[index]['start_time']  ,"%Y-%m-%d %H:%M:%S")
                        except:
                            init_time = datetime.datetime.strptime(df[index]['start_time']  ,"%Y-%m-%d %H:%M:%S.%f")               
                    else:            
                        try:
                            end_time =  datetime.datetime.strptime(df[index]['start_time']  ,"%Y-%m-%d %H:%M:%S")
                        except: # 會這樣寫是 因為  csv 每個  start time 有可能會遇到 是時間戳的  ,或者  excle最後一筆 是沒有 start time的
                            end_time =  datetime.datetime.strptime(df[index]['start_time']  ,"%Y-%m-%d %H:%M:%S.%f")
                
            except Exception as e:
                print('df_dict: %s , error: %s'%(df_dict , e))
                pass
        
        self.all_time = (end_time - init_time).total_seconds() #這個會拿來 用  每個 API 所有次數  /  執行時間
        self.all_time_title = '執行時間: %s ~ %s , 秒數: %s '%(init_time ,end_time ,  self.all_time)
        print('all_time: %s'%self.all_time)        
        # print('statistic_dict: %s'%self.statistic_dict )

    def cal_num_inter(self , test_num):# 針對 紅包雨 的reponse 時間 , 產出是屬於哪個區間 邏輯
        #print('test_num: %s'%test_num)
        if test_num < 0:
            test_num = test_num + 120        
        if test_num >0 and test_num < 1:
            format_interval = "0 - 1秒"
        elif test_num >1 and test_num < 2:
            format_interval = "1 - 2秒"
        elif test_num > 2 and test_num < 3:
            format_interval = "2 - 3秒"
        elif test_num < 5:
            format_interval = "3 - 5秒"
        elif test_num <=10 :
            format_interval = "5 ~ 10秒"
        elif test_num >= 360:
            format_interval = "360秒 以上"
     
        else: # 計算區間 ,
            if test_num < 60:# 小60  , 每5 為區間
                '''
                將數值 除以10 後 ,  拆分成 整數 和 小數點
                ex: 23.89 會轉成 23 後 ,  再拆成 整數 2 和 小數點 0.3
                並用 小數點 來盤斷是否大於 0.5 
                '''
                devide_num = 10

            else:# 大於60  ,每 30 為區間
                devide_num = 60

                '''
                將數值 除以60 後 ,  拆分  整數 和 小數點
                ex: 96.89 會轉成 96 後 ,  再拆成 整數 1 和 小數點 0.6
                並用 小數點 來盤斷是否大於 0.5 
                '''

            t_division = test_num  / devide_num
            #print(t_division)
            t_int = int(t_division)   # 整數
            t_float =  float(t_division - t_int ) # 小數點
            #print('整數 : %s , 小數: %s'%(t_int , t_float ) )
            
            #ex : 23.81 :  t_int為2 , t_float為 0.3 屬於 20 - 25
            #ex: 90.12 : t_int為 1 , t_float為 0.5 屬於 90 - 120
            if t_float  >= 0.5:# 需進位
                low_string = int( (t_int + 0.5)    * devide_num)
                max_string = int((t_int + 1 ) * devide_num )
            else:#
                low_string = int(t_int * devide_num)
                max_string = int( (t_int + 0.5 ) * devide_num )

            format_interval = "{low_string} ~ {max_string} 秒".format(low_string = low_string , 
                max_string = max_string )
        return format_interval

    def generate_data(self , sheet_name):# sheet data邏輯

        if sheet_name == 'statistic':
            list_without_nan = []
            
            remove_diff_nan_list = [] # 存放 沒有 diff 的 action
            for key in list(self.statistic_dict.keys()):
                try:
                    if type(key) is  str:
                        list_without_nan.append(key)
                except: # nan
                    pass

            frame_data  = {
                "動作": list_without_nan ,
                "次數": [] ,
                "成功": [],
                "成功率": [],
                "Avg.Response Time":[] , 
                "Throughput(APIs / sec)":[]
                }
            
            for action in list_without_nan:          
                try:
                    if 'WS_Error' in action:
                        error_times = len(self.statistic_dict[action]['error_times'] )
                        frame_data['次數'].append(error_times)
                        frame_data['成功'].append('')
                        frame_data['成功率'].append('')
                        frame_data['Avg.Response Time'].append('')
                        frame_data['Throughput(APIs / sec)'].append('')                   
                    else:
                        status_code_list =  self.statistic_dict[action]['status_code']
                        diff_list = self.statistic_dict[action]['diff']
                        if not diff_list: #空
                            remove_diff_nan_list.append(action)
                            continue

                        action_time = len(status_code_list)# 該action 的總次數
                        frame_data['次數'].append(action_time)

                        if action == 'Rain':
                            assert_code = '0'
                        else:
                            assert_code = '200'

                        success_time =  len(list(filter(lambda score: score == assert_code, status_code_list ) ) )
                        frame_data['成功'].append(success_time)

                        success_percent = '{percent}%'.format(percent = round( success_time/action_time * 100 ,2 ) )
                        frame_data['成功率'].append(success_percent)

                        ava_response = round(reduce(lambda a, b: a + b, diff_list ) / len(diff_list) ,2)
                        frame_data['Avg.Response Time'].append(ava_response)

                        Throughput  = round(action_time / self.all_time , 2) # 平均每秒  多少個api 請求 :   所有次數 / 整個測項 執行多少秒
                        frame_data['Throughput(APIs / sec)'].append(Throughput )
               
                except Exception as e:
                    print('action: %s error: %s'%( action , e ))
                    
                    frame_data['次數'].append('')
                    frame_data['成功'].append('')
                    frame_data['成功率'].append('')
                    frame_data['Avg.Response Time'].append('')
                    frame_data['Throughput(APIs / sec)'].append('')
            
            for no_diff_action in remove_diff_nan_list:
                frame_data['動作'].remove(no_diff_action)
            #print(frame_data)

        elif sheet_name == 'socket 延遲時間':
            len_rain_success = len( self.socket_rain_success )
            #print(' 紅包雨: %s'%self.socket_rain_success)

            frame_data  = {
                "級距": [],
                "數量": [] ,
                "比例": [],
                }
            socket_deafault = defaultdict(list)# 用來存放  key為 區間名稱 , value為list存放 數量
            
            # 先將所有 紅包雨 成功的 時間 loop 出來 ,確認 每個 區間名稱 數量
            for rain_time in sorted(self.socket_rain_success):
                interval_key = self.cal_num_inter(test_num= rain_time)# 紅包雨 時間 產出 屬於哪個區間的 key
                socket_deafault[interval_key].append(1)
            
            for interval_key in socket_deafault.keys():#interval_key 在這都會是唯一值
                frame_data['級距'].append(interval_key)
                
                rain_num = len( socket_deafault[interval_key] )
                frame_data['數量'].append(  rain_num )

                rain_percent = '{percent}%'.format(percent = round( rain_num/len_rain_success * 100 ,2 ) )
                frame_data['比例'].append(  rain_percent )       
        else:
            return False

        return frame_data



'''
Xlsx
csv_filter_name 參數預設 為qa (會找 config 路徑下的所有 相關 qa csv檔)
'''
xlsx = Xlsx( csv_filter_name = 'async' ) # qa  , async   , docker 
if len(xlsx.csv_list) == 0:
    print('找不到相關csv 檔案 ,　請確認config csv_filename_path 路徑')

else:
    print(xlsx.csv_list)
    for index ,  filter_csv in enumerate(xlsx.csv_list):
        print('filter_csv: %s'%filter_csv)
        try:
            if index == 0:
                xlsx.csv_to_xlsx(csv_name = filter_csv )
            else:
                xlsx.update_sheet_1(csv_name = filter_csv )
        except Exception as e:
            print('error: %s'%e)
            pass
#進球紅包雨，才需增加goal_event_tiem欄位
if xlsx.test_type == 'goal':
    xlsx.add_column_to_sheet_1()
xlsx.xlsx_data_transfer()#把需要的 csv 都轉完成 xlsx後 , 就可以 讀取 xlsx, 並轉成 需要的字典
xlsx.add_sheet_data(sheet_name = 'statistic' )# 各action 時間 統計
xlsx.add_sheet_data(sheet_name = 'socket 延遲時間' )# 針對 Rain 的成功 時間 做 統計 , 如果 action 沒有 Rain　將會沒有資料


for i in xlsx.rain_user_list.keys():
    
    diff_list = list(set(xlsx.all_user ).difference(set(xlsx.rain_user_list[i] )))
    print( '%s 沒領取rain 的 user:%s , len :%s '%(i , diff_list  , len(diff_list) )    )